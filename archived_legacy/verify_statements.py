import os
import re
import json
import time
import math
import hashlib
import argparse
from dataclasses import dataclass
from typing import List, Dict, Tuple, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

import numpy as np
import pandas as pd
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type
from tqdm import tqdm

from openpyxl import Workbook, load_workbook
from openai import OpenAI


# -----------------------------
# Configuration data classes
# -----------------------------

@dataclass
class IndexMetadata:
	corpus_path: str
	file_size: int
	mtime: float
	chunk_size_chars: int
	overlap_chars: int
	embedding_model: str


# -----------------------------
# Utilities
# -----------------------------

def load_api_credentials(api_file_path: str) -> Tuple[str, Optional[str]]:
	"""Parse a simple key=value file and return (api_key, base_url)."""
	api_key = None
	base_url = None
	if not os.path.exists(api_file_path):
		raise FileNotFoundError(f"API file not found: {api_file_path}")
	with open(api_file_path, "r", encoding="utf-8") as f:
		for line in f:
			line = line.strip()
			if not line or line.startswith("#"):
				continue
			if "=" in line:
				k, v = line.split("=", 1)
				k = k.strip()
				v = v.strip()
				# remove inline comments and surrounding quotes
				if "#" in v:
					v = v.split("#", 1)[0].strip()
				if (v.startswith('"') and v.endswith('"')) or (v.startswith("'") and v.endswith("'")):
					v = v[1:-1]
				if k in ("OPENAI_API_KEY", "API_KEY") and v:
					api_key = v
				elif k in ("OPENAI_BASE_URL", "BASE_URL") and v:
					base_url = v
	# Fallback to environment variables
	if not api_key:
		api_key = os.getenv("OPENAI_API_KEY") or os.getenv("API_KEY")
	if not base_url:
		base_url = os.getenv("OPENAI_BASE_URL") or os.getenv("BASE_URL")
	if not api_key:
		raise RuntimeError(f"OPENAI_API_KEY not found in: {api_file_path}. Add 'OPENAI_API_KEY=sk-...' (no quotes) or set env var OPENAI_API_KEY.")
	# Normalize base_url for OpenAI public endpoint to ensure /v1 is present
	if base_url:
		url = base_url.strip()
		if url.startswith("https://api.openai.com") and not url.rstrip("/").endswith("/v1"):
			base_url = url.rstrip("/") + "/v1"
	return api_key, base_url


def fingerprint_file(path: str) -> Tuple[int, float]:
	stat = os.stat(path)
	return stat.st_size, stat.st_mtime


def compute_index_id(meta: IndexMetadata) -> str:
	payload = json.dumps({
		"corpus_path": os.path.basename(meta.corpus_path),
		"file_size": meta.file_size,
		"mtime": meta.mtime,
		"chunk_size_chars": meta.chunk_size_chars,
		"overlap_chars": meta.overlap_chars,
		"embedding_model": meta.embedding_model,
	}, sort_keys=True).encode("utf-8")
	return hashlib.sha1(payload).hexdigest()


def ensure_dir(path: str) -> None:
	os.makedirs(path, exist_ok=True)


# -----------------------------
# Chunking and embeddings
# -----------------------------

def chunk_text(text: str, chunk_size_chars: int, overlap_chars: int) -> List[Tuple[int, str]]:
	"""Return list of (start_index, chunk_text)."""
	chunks: List[Tuple[int, str]] = []
	start = 0
	length = len(text)
	if chunk_size_chars <= 0:
		raise ValueError("chunk_size_chars must be > 0")
	if overlap_chars < 0:
		raise ValueError("overlap_chars must be >= 0")
	while start < length:
		end = min(length, start + chunk_size_chars)
		chunk = text[start:end]
		chunks.append((start, chunk))
		if end == length:
			break
		start = end - overlap_chars if end - overlap_chars > start else end
	return chunks


@retry(stop=stop_after_attempt(6), wait=wait_exponential(multiplier=1, min=1, max=30), reraise=True)
def embed_texts(client: OpenAI, model: str, inputs: List[str]) -> np.ndarray:
	# Increased batch size for better performance
	batch_size = 256  # Increased from 128
	all_vectors = []
	
	for i in range(0, len(inputs), batch_size):
		batch = inputs[i:i + batch_size]
		resp = client.embeddings.create(model=model, input=batch)
		vectors = [np.array(d.embedding, dtype=np.float32) for d in resp.data]
		all_vectors.extend(vectors)
	
	return np.vstack(all_vectors)


def extract_text_from_pdf(pdf_path: str) -> str:
	"""Extract text from PDF file."""
	try:
		import PyPDF2
		text = ""
		with open(pdf_path, 'rb') as file:
			pdf_reader = PyPDF2.PdfReader(file)
			for page_num in range(len(pdf_reader.pages)):
				page = pdf_reader.pages[page_num]
				text += page.extract_text() + "\n"
		return text
	except ImportError:
		raise RuntimeError("PyPDF2 not installed. Run: pip install PyPDF2")
	except Exception as e:
		raise RuntimeError(f"Error reading PDF {pdf_path}: {str(e)}")

def extract_text_from_word(docx_path: str) -> str:
	"""Extract text from Word document."""
	try:
		from docx import Document
		doc = Document(docx_path)
		text = ""
		for paragraph in doc.paragraphs:
			text += paragraph.text + "\n"
		return text
	except ImportError:
		raise RuntimeError("python-docx not installed. Run: pip install python-docx")
	except Exception as e:
		raise RuntimeError(f"Error reading Word document {docx_path}: {str(e)}")

def extract_text_from_excel(excel_path: str) -> str:
	"""Extract text from Excel file by reading all sheets."""
	try:
		# Read all sheets
		excel_file = pd.ExcelFile(excel_path)
		text = ""
		
		for sheet_name in excel_file.sheet_names:
			df = pd.read_excel(excel_path, sheet_name=sheet_name)
			text += f"\n--- Sheet: {sheet_name} ---\n"
			
			# Convert all non-null values to text
			for col in df.columns:
				for value in df[col].dropna():
					text += str(value) + "\n"
		
		return text
	except Exception as e:
		raise RuntimeError(f"Error reading Excel file {excel_path}: {str(e)}")

def read_source_file(file_path: str) -> str:
	"""Read text from various file formats."""
	file_ext = os.path.splitext(file_path)[1].lower()
	
	if file_ext == '.txt':
		with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
			return f.read()
	elif file_ext == '.pdf':
		return extract_text_from_pdf(file_path)
	elif file_ext in ['.docx', '.doc']:
		return extract_text_from_word(file_path)
	elif file_ext in ['.xlsx', '.xls']:
		return extract_text_from_excel(file_path)
	else:
		raise RuntimeError(f"Unsupported file format: {file_ext}. Supported formats: .txt, .pdf, .docx, .doc, .xlsx, .xls")


def build_or_load_index(corpus_path: str,
						 embedding_model: str,
						 client: OpenAI,
						 chunk_size_chars: int,
						 overlap_chars: int,
						 cache_dir: str = ".index_cache") -> Tuple[np.ndarray, List[Tuple[int, str]], str]:
	"""
	Build or load an embeddings index. Returns (embeddings_matrix, chunks, index_id).
	"""
	file_size, mtime = fingerprint_file(corpus_path)
	meta = IndexMetadata(
		corpus_path=corpus_path,
		file_size=file_size,
		mtime=mtime,
		chunk_size_chars=chunk_size_chars,
		overlap_chars=overlap_chars,
		embedding_model=embedding_model,
	)
	index_id = compute_index_id(meta)
	ensure_dir(cache_dir)
	emb_path = os.path.join(cache_dir, f"{index_id}.npz")
	chunks_path = os.path.join(cache_dir, f"{index_id}.chunks.jsonl")

	if os.path.exists(emb_path) and os.path.exists(chunks_path):
		data = np.load(emb_path)
		embeddings = data["embeddings"]
		chunks: List[Tuple[int, str]] = []
		with open(chunks_path, "r", encoding="utf-8") as f:
			for line in f:
				obj = json.loads(line)
				chunks.append((obj["start"], obj["text"]))
		return embeddings, chunks, index_id

	# Build fresh - now supports multiple file formats
	print(f"Reading source file: {corpus_path}")
	text = read_source_file(corpus_path)
	print(f"Extracted {len(text)} characters from source file")
	
	chunks = chunk_text(text, chunk_size_chars, overlap_chars)

	# Batch embed with larger batch size for better performance
	batch_size = 256  # Increased from 128
	all_vecs: List[np.ndarray] = []
	print(f"Embedding {len(chunks)} chunks in batches of {batch_size}...")
	for i in tqdm(range(0, len(chunks), batch_size), desc="Embedding corpus", unit="batch"):
		batch = [c[1] for c in chunks[i:i+batch_size]]
		vecs = embed_texts(client, embedding_model, batch)
		all_vecs.append(vecs)
	embeddings = np.vstack(all_vecs)

	# Save cache with metadata
	np.savez_compressed(emb_path, embeddings=embeddings)
	with open(chunks_path, "w", encoding="utf-8") as f:
		# Write metadata first
		f.write(json.dumps({
			"metadata": {
				"corpus_path": os.path.basename(corpus_path),
				"file_size": file_size,
				"mtime": mtime,
				"chunk_size_chars": chunk_size_chars,
				"overlap_chars": overlap_chars,
				"embedding_model": embedding_model,
			}
		}, ensure_ascii=False) + "\n")
		# Write chunks
		for start, ctext in chunks:
			f.write(json.dumps({"start": start, "text": ctext}, ensure_ascii=False) + "\n")
	return embeddings, chunks, index_id


def normalize_matrix(matrix: np.ndarray) -> np.ndarray:
	norms = np.linalg.norm(matrix, axis=1, keepdims=True)
	norms[norms == 0] = 1.0
	return matrix / norms


def top_k_indices_cosine(doc_matrix_norm: np.ndarray, query_vec: np.ndarray, k: int) -> List[int]:
	q = query_vec / (np.linalg.norm(query_vec) + 1e-12)
	scores = doc_matrix_norm @ q
	if k >= len(scores):
		return list(np.argsort(-scores))
	# Partial sort then final sort for stability
	idx = np.argpartition(-scores, k)[:k]
	idx = idx[np.argsort(-scores[idx])]
	return idx.tolist()


# -----------------------------
# LLM Verification
# -----------------------------

def build_verification_prompt(statement: str, evidence_snippets: List[str]) -> List[Dict[str, str]]:
	snippets_joined = "\n\n--- Evidence Snippet ---\n\n".join(evidence_snippets)
	system = (
		"You are an experienced legal council and fact-checker. Given a STATEMENT and EVIDENCE snippets from a source document, "
		"decide if the statement is accurate according to the evidence. If evidence is insufficient or contradicts, "
		"return a 'not accurate' verdict and explain briefly using references to the snippets. Output strict JSON."
	)
	user = (
		f"STATEMENT:\n{statement}\n\n"
		f"EVIDENCE SNIPPETS:\n{snippets_joined}\n\n"
		"Return a JSON object with keys: verdict ('accurate'|'not accurate'), explanation (string)."
	)
	return [
		{"role": "system", "content": system},
		{"role": "user", "content": user},
	]

def build_court_verification_prompt(par_number: str, three_line_content: str, evidence_snippets: List[str]) -> List[Dict[str, str]]:
	snippets_joined = "\n\n--- Evidence Snippet ---\n\n".join(evidence_snippets)
	system = (
		"You are an experienced legal council and fact-checker analyzing court judgment accuracy. Given a PARAGRAPH from a court judgment "
		"and EVIDENCE snippets from source documents, determine if the paragraph accurately represents what is said in the evidence. "
		"If accurate, return 'accurate'. If not accurate, identify the type of inaccuracy and provide a brief explanation. "
		"Also provide a degree of accuracy on a scale from 1 to 10. "
		"Output strict JSON."
	)
	user = (
		f"PARAGRAPH {par_number}:\n{three_line_content}\n\n"
		f"EVIDENCE SNIPPETS:\n{snippets_joined}\n\n"
		"Return a JSON object with keys: "
		"verdict ('accurate'|'not accurate'), "
		"degree_of_accuracy (integer 1-10, where 10 is completely accurate), "
		"inaccuracy_type ('manipulation'|'omission'|'addition'|'none'), "
		"description (25 words max explaining the inaccuracy if any)."
	)
	return [
		{"role": "system", "content": system},
		{"role": "user", "content": user},
	]

@retry(stop=stop_after_attempt(6), wait=wait_exponential(multiplier=1, min=1, max=30), reraise=True,
	   retry=retry_if_exception_type(Exception))
def judge_statement(client: OpenAI, model: str, statement: str, evidence_snippets: List[str]) -> Dict[str, str]:
	messages = build_verification_prompt(statement, evidence_snippets)
	resp = client.chat.completions.create(
		model=model,
		messages=messages,
		response_format={"type": "json_object"},
	)
	content = resp.choices[0].message.content
	try:
		obj = json.loads(content)
		verdict = obj.get("verdict", "not accurate").strip().lower()
		explanation = obj.get("explanation", "")
		if verdict not in ("accurate", "not accurate"):
			verdict = "not accurate"
		return {"verdict": verdict, "explanation": explanation}
	except Exception:
		# Fallback minimal parsing
		text = content.strip()
		if "accurate" in text.lower() and "not" not in text.lower():
			return {"verdict": "accurate", "explanation": text}
		return {"verdict": "not accurate", "explanation": text}

@retry(stop=stop_after_attempt(6), wait=wait_exponential(multiplier=1, min=1, max=30), reraise=True,
	   retry=retry_if_exception_type(Exception))
def judge_court_statement(client: OpenAI, model: str, par_number: str, three_line_content: str, evidence_snippets: List[str]) -> Dict[str, str]:
	messages = build_court_verification_prompt(par_number, three_line_content, evidence_snippets)
	resp = client.chat.completions.create(
		model=model,
		messages=messages,
		response_format={"type": "json_object"},
		timeout=30,  # 30 second timeout
	)
	content = resp.choices[0].message.content
	try:
		obj = json.loads(content)
		verdict = obj.get("verdict", "not accurate").strip().lower()
		degree_of_accuracy = obj.get("degree_of_accuracy", 5)  # Default to 5 if not provided
		inaccuracy_type = obj.get("inaccuracy_type", "none").strip().lower()
		description = obj.get("description", "")
		
		if verdict not in ("accurate", "not accurate"):
			verdict = "not accurate"
		if inaccuracy_type not in ("manipulation", "omission", "addition", "none"):
			inaccuracy_type = "none"
		
		# Ensure degree_of_accuracy is a valid integer between 1-10
		try:
			degree_of_accuracy = int(degree_of_accuracy)
			if degree_of_accuracy < 1:
				degree_of_accuracy = 1
			elif degree_of_accuracy > 10:
				degree_of_accuracy = 10
		except (ValueError, TypeError):
			degree_of_accuracy = 5  # Default if parsing fails
		
		return {
			"verdict": verdict, 
			"degree_of_accuracy": degree_of_accuracy,
			"inaccuracy_type": inaccuracy_type if verdict == "not accurate" else "none",
			"description": description
		}
	except Exception:
		# Fallback minimal parsing
		text = content.strip()
		if "accurate" in text.lower() and "not" not in text.lower():
			return {"verdict": "accurate", "degree_of_accuracy": 10, "inaccuracy_type": "none", "description": text}
		return {"verdict": "not accurate", "degree_of_accuracy": 3, "inaccuracy_type": "manipulation", "description": text}


# -----------------------------
# I/O helpers
# -----------------------------

def read_statements(path: str) -> List[str]:
	text: List[str] = []
	if path.lower().endswith(".txt"):
		with open(path, "r", encoding="utf-8", errors="ignore") as f:
			for line in f:
				line = line.strip()
				if line:
					text.append(line)
	elif path.lower().endswith((".xlsx", ".xls")):
		df = pd.read_excel(path)
		# Look for specific columns for court judgment format
		par_col = None
		content_col = None
		
		# Find Par and Content columns (case insensitive)
		for c in df.columns:
			col_lower = str(c).strip().lower()
			if col_lower == "par":
				par_col = c
			elif col_lower == "content":
				content_col = c
		
		if par_col is None or content_col is None:
			# Fallback to heuristics if specific columns not found
			for c in df.columns:
				if str(c).strip().lower() in ("statement", "statements", "claim", "text"):
					content_col = c
					break
			if content_col is None:
				for c in df.columns:
					if df[c].notna().any():
						content_col = c
						break
			if content_col is None:
				raise RuntimeError("No 'Content' column found in the Excel statements file")
			
			# For non-court format, just return content
			text = [str(v).strip() for v in df[content_col].tolist() if pd.notna(v) and str(v).strip()]
		else:
			# Court judgment format - return structured data
			text = []
			for idx, row in df.iterrows():
				if pd.notna(row[content_col]) and str(row[content_col]).strip():
					text.append({
						'par': str(row[par_col]).strip() if pd.notna(row[par_col]) else str(idx + 1),
						'content': str(row[content_col]).strip(),
						'index': idx
					})
	else:
		raise RuntimeError("Unsupported statements file format. Use .txt or .xlsx")
	return text


def excel_prepare_writer(path: str, headers: List[str]) -> Tuple[str, Workbook]:
	# Always create new workbook or overwrite existing one
	wb = Workbook()
	ws = wb.active
	ws.title = "Results"
	ws.append(headers)
	wb.save(path)
	return path, wb

def excel_prepare_court_writer(path: str) -> Tuple[str, Workbook]:
	headers = ["Analysed Par number", "Par Context", "Is Accurate", "Degree of Accuracy", "Inaccuracy Type", "Description"]
	# Always create new workbook or overwrite existing one
	wb = Workbook()
	ws = wb.active
	ws.title = "Court Analysis Results"
	ws.append(headers)
	
	# Try to save with error handling
	try:
		wb.save(path)
	except PermissionError:
		# If file is in use, try with a timestamp
		import datetime
		timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
		base_name = os.path.splitext(path)[0]
		ext = os.path.splitext(path)[1]
		new_path = f"{base_name}_{timestamp}{ext}"
		print(f"Warning: {path} is in use. Saving to {new_path} instead.")
		wb.save(new_path)
		return new_path, wb
	
	return path, wb


def excel_append_row(path: str, row: List[str]) -> None:
	try:
		wb = load_workbook(path)
		ws = wb.active
		ws.append(row)
		wb.save(path)
	except PermissionError:
		# If file is in use, try with a timestamp
		import datetime
		timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
		base_name = os.path.splitext(path)[0]
		ext = os.path.splitext(path)[1]
		new_path = f"{base_name}_{timestamp}{ext}"
		print(f"Warning: {path} is in use. Saving to {new_path} instead.")
		wb.save(new_path)
		# Update the global path for future saves
		global current_excel_path
		current_excel_path = new_path


def sort_excel_by_paragraph_number(path: str) -> None:
	"""Sort the Excel file by paragraph number (first column)."""
	try:
		# Read the Excel file
		df = pd.read_excel(path)
		
		# Convert paragraph numbers to numeric for proper sorting
		# Handle both numeric and string paragraph numbers
		def extract_number(par_str):
			try:
				# Try to extract numeric part from string
				import re
				numbers = re.findall(r'\d+', str(par_str))
				if numbers:
					return int(numbers[0])
				return 0
			except:
				return 0
		
		# Create a sorting key from the first column
		df['sort_key'] = df.iloc[:, 0].apply(extract_number)
		
		# Sort by the numeric key
		df_sorted = df.sort_values('sort_key')
		
		# Remove the temporary sort key column
		df_sorted = df_sorted.drop('sort_key', axis=1)
		
		# Write back to Excel with error handling
		try:
			df_sorted.to_excel(path, index=False)
			print(f"Sorted {path} by paragraph number")
		except PermissionError:
			# If file is in use, try with a timestamp
			import datetime
			timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
			base_name = os.path.splitext(path)[0]
			ext = os.path.splitext(path)[1]
			new_path = f"{base_name}_{timestamp}{ext}"
			print(f"Warning: {path} is in use. Saving sorted results to {new_path} instead.")
			df_sorted.to_excel(new_path, index=False)
			print(f"Sorted {new_path} by paragraph number")
		
	except Exception as e:
		print(f"Warning: Could not sort Excel file: {str(e)}")


def load_court_checkpoint(path: str) -> Dict[str, Dict[str, str]]:
	results: Dict[str, Dict[str, str]] = {}
	if os.path.exists(path):
		with open(path, "r", encoding="utf-8") as f:
			for line in f:
				try:
					obj = json.loads(line)
					# Handle both old and new checkpoint formats
					par_key = obj.get("par_number") or obj.get("index")
					if par_key:
						results[str(par_key)] = obj
				except (json.JSONDecodeError, KeyError):
					# Skip malformed lines
					continue
	return results

def load_checkpoint(path: str) -> Dict[int, Dict[str, str]]:
	results: Dict[int, Dict[str, str]] = {}
	if os.path.exists(path):
		with open(path, "r", encoding="utf-8") as f:
			for line in f:
				try:
					obj = json.loads(line)
					# Handle both old and new checkpoint formats
					idx_key = obj.get("index") or obj.get("par_number")
					if idx_key is not None:
						results[int(idx_key)] = obj
				except (json.JSONDecodeError, KeyError, ValueError):
					# Skip malformed lines
					continue
	return results


def append_checkpoint(path: str, index: int, statement: str, verdict: str, explanation: str) -> None:
	with open(path, "a", encoding="utf-8") as f:
		f.write(json.dumps({
			"index": index,
			"statement": statement,
			"verdict": verdict,
			"explanation": explanation,
		}, ensure_ascii=False) + "\n")

def append_court_checkpoint(path: str, par_number: str, par_context: str, verdict: str, degree_of_accuracy: int, inaccuracy_type: str, description: str) -> None:
	with open(path, "a", encoding="utf-8") as f:
		f.write(json.dumps({
			"par_number": par_number,
			"par_context": par_context,
			"verdict": verdict,
			"degree_of_accuracy": degree_of_accuracy,
			"inaccuracy_type": inaccuracy_type,
			"description": description,
		}, ensure_ascii=False) + "\n")


# -----------------------------
# Main pipeline
# -----------------------------

def get_file_size_mb(file_path: str) -> float:
	"""Get file size in MB."""
	if os.path.exists(file_path):
		return os.path.getsize(file_path) / (1024 * 1024)
	return 0.0

def list_available_caches(cache_dir: str = ".index_cache") -> List[Dict[str, str]]:
	"""List all available cached files with metadata."""
	caches = []
	if not os.path.exists(cache_dir):
		return caches
	
	for file in os.listdir(cache_dir):
		if file.endswith(".npz"):
			cache_id = file[:-4]  # Remove .npz
			chunks_file = os.path.join(cache_dir, f"{cache_id}.chunks.jsonl")
			
			if os.path.exists(chunks_file):
				# Try to extract original filename from chunks
				original_name = cache_id
				try:
					with open(chunks_file, "r", encoding="utf-8") as f:
						first_line = f.readline()
						if first_line:
							obj = json.loads(first_line)
							if "metadata" in obj:
								original_name = obj["metadata"].get("corpus_path", cache_id)
				except:
					pass
				
				# Calculate cache sizes
				npz_size = get_file_size_mb(os.path.join(cache_dir, file))
				chunks_size = get_file_size_mb(chunks_file)
				total_size = npz_size + chunks_size
				
				caches.append({
					"cache_id": cache_id,
					"original_name": original_name,
					"npz_path": os.path.join(cache_dir, file),
					"chunks_path": chunks_file,
					"npz_size_mb": npz_size,
					"chunks_size_mb": chunks_size,
					"total_size_mb": total_size
				})
	
	return caches

def load_multiple_caches(cache_ids: List[str], cache_dir: str = ".index_cache") -> Tuple[np.ndarray, List[Tuple[int, str]], List[str]]:
	"""Load multiple caches and combine them."""
	all_embeddings = []
	all_chunks = []
	cache_names = []
	total_size = 0.0
	
	for cache_id in cache_ids:
		npz_path = os.path.join(cache_dir, f"{cache_id}.npz")
		chunks_path = os.path.join(cache_dir, f"{cache_id}.chunks.jsonl")
		
		if not (os.path.exists(npz_path) and os.path.exists(chunks_path)):
			print(f"Warning: Cache {cache_id} not found, skipping")
			continue
		
		# Calculate cache size
		cache_size = get_file_size_mb(npz_path) + get_file_size_mb(chunks_path)
		total_size += cache_size
		
		# Load embeddings
		data = np.load(npz_path)
		embeddings = data["embeddings"]
		
		# Load chunks
		chunks = []
		with open(chunks_path, "r", encoding="utf-8") as f:
			for line in f:
				obj = json.loads(line)
				# Skip metadata line (first line with "metadata" key)
				if "metadata" in obj:
					continue
				# Only process lines with "start" and "text" keys
				if "start" in obj and "text" in obj:
					chunks.append((obj["start"], obj["text"]))
		
		# Get original filename
		original_name = cache_id
		try:
			with open(chunks_path, "r", encoding="utf-8") as f:
				first_line = f.readline()
				if first_line:
					obj = json.loads(first_line)
					if "metadata" in obj:
						original_name = obj["metadata"].get("corpus_path", cache_id)
		except:
			pass
		
		all_embeddings.append(embeddings)
		all_chunks.extend(chunks)
		cache_names.append(original_name)
	
	if not all_embeddings:
		raise RuntimeError("No valid caches found")
	
	# Combine all embeddings
	combined_embeddings = np.vstack(all_embeddings)
	
	print(f"Loaded {len(cache_ids)} caches, total size: {total_size:.1f} MB")
	
	return combined_embeddings, all_chunks, cache_names

def select_caches_gui() -> Tuple[List[str], str]:
	"""GUI to select which caches to use. Returns (selected_cache_ids, statements_file_path)."""
	try:
		import tkinter as tk
		from tkinter import ttk, messagebox, filedialog
	except ImportError:
		print("tkinter not available, using command line selection")
		caches = select_caches_cli()
		return caches, None
	
	caches = list_available_caches()
	
	root = tk.Tk()
	root.title("Select Source Caches")
	root.geometry("800x500")
	
	# Create main frame
	main_frame = ttk.Frame(root, padding="10")
	main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
	
	# Instructions
	ttk.Label(main_frame, text="Select which cached source files to use for analysis:", font=("Arial", 10, "bold")).grid(row=0, column=0, columnspan=3, pady=(0, 10), sticky=tk.W)
	
	# Create treeview for better display
	tree_frame = ttk.Frame(main_frame)
	tree_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
	
	# Create treeview with columns
	columns = ("filename", "size", "id")
	tree = ttk.Treeview(tree_frame, columns=columns, show="tree headings", height=15)
	
	# Configure columns
	tree.heading("filename", text="Source File")
	tree.heading("size", text="Cache Size (MB)")
	tree.heading("id", text="Cache ID")
	
	tree.column("filename", width=300, anchor=tk.W)
	tree.column("size", width=100, anchor=tk.E)
	tree.column("id", width=150, anchor=tk.W)
	
	# Add scrollbar
	scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
	tree.configure(yscrollcommand=scrollbar.set)
	
	tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
	scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
	
	# Populate treeview
	for cache in caches:
		tree.insert("", tk.END, values=(
			cache['original_name'],
			f"{cache['total_size_mb']:.1f}",
			cache['cache_id']
		))
	
	# Select all by default
	for item in tree.get_children():
		tree.selection_add(item)
	
	# Total size display
	total_size = sum(cache['total_size_mb'] for cache in caches)
	size_label = ttk.Label(main_frame, text=f"Total cache size: {total_size:.1f} MB", font=("Arial", 9))
	size_label.grid(row=2, column=0, columnspan=2, pady=(5, 0), sticky=tk.W)
	
	# Buttons
	button_frame = ttk.Frame(main_frame)
	button_frame.grid(row=3, column=0, columnspan=3, pady=(10, 0))
	
	selected_caches = []
	selected_statements_file = None
	
	def update_total_size():
		"""Update total size display based on selection."""
		selection = tree.selection()
		selected_size = sum(
			cache['total_size_mb'] 
			for cache in caches 
			if cache['cache_id'] in [tree.item(item)['values'][2] for item in selection]
		)
		size_label.config(text=f"Selected cache size: {selected_size:.1f} MB")
	
	def on_ok():
		nonlocal selected_caches
		selection = tree.selection()
		if not selection:
			messagebox.showwarning("No Selection", "Please select at least one cache.")
			return
		
		if not selected_statements_file:
			messagebox.showwarning("No Statements File", "Please select an Excel file with paragraphs to verify.")
			return
		
		selected_caches = [tree.item(item)['values'][2] for item in selection]
		root.destroy()
	
	def on_cancel():
		root.destroy()
	
	def on_select_all():
		for item in tree.get_children():
			tree.selection_add(item)
		update_total_size()
	
	def on_deselect_all():
		tree.selection_remove(tree.selection())
		update_total_size()
	
	def on_delete_selected_caches():
		"""Delete selected caches from the main GUI."""
		nonlocal caches
		
		selection = tree.selection()
		if not selection:
			messagebox.showwarning("No Selection", "Please select at least one cache to delete.")
			return
		
		selected_caches = [tree.item(item)['values'][2] for item in selection]
		selected_names = [tree.item(item)['values'][0] for item in selection]
		selected_size = sum(
			cache['total_size_mb'] 
			for cache in caches 
			if cache['cache_id'] in selected_caches
		)
		
		# Confirmation dialog
		confirm_text = f"Are you sure you want to delete {len(selected_caches)} cache(s)?\n\n"
		confirm_text += f"Files to delete:\n"
		for name in selected_names:
			confirm_text += f"• {name}\n"
		confirm_text += f"\nTotal size to free: {selected_size:.1f} MB\n\nThis action cannot be undone!"
		
		if not messagebox.askyesno("Confirm Deletion", confirm_text):
			return
		
		# Delete selected caches
		success_count = 0
		for cache_id in selected_caches:
			if delete_cache(cache_id):
				success_count += 1
		
		# Show results
		if success_count == len(selected_caches):
			messagebox.showinfo("Success", f"Successfully deleted {success_count} cache(s).")
		else:
			messagebox.showwarning("Partial Success", f"Deleted {success_count}/{len(selected_caches)} cache(s). Some files may have been in use.")
		
		# Refresh the cache list
		caches = list_available_caches()
		
		# Clear and repopulate treeview
		for item in tree.get_children():
			tree.delete(item)
		
		for cache in caches:
			tree.insert("", tk.END, values=(
				cache['original_name'],
				f"{cache['total_size_mb']:.1f}",
				cache['cache_id']
			))
		
		# Select all by default
		for item in tree.get_children():
			tree.selection_add(item)
		
		# Update total size
		total_size = sum(cache['total_size_mb'] for cache in caches)
		size_label.config(text=f"Total cache size: {total_size:.1f} MB")
	
	def on_add_source_file():
		"""Add a new source file and refresh the cache list."""
		# Select source file
		source_file = filedialog.askopenfilename(
			title="Select source file (TXT, PDF, Word, Excel)",
			filetypes=[
				("All supported files", "*.txt *.pdf *.docx *.doc *.xlsx *.xls"),
				("Text files", "*.txt"),
				("PDF files", "*.pdf"),
				("Word documents", "*.docx *.doc"),
				("Excel files", "*.xlsx *.xls"),
				("All files", "*.*")
			]
		)
		
		if not source_file:
			return
		
		try:
			# Load API credentials
			api_key, base_url = load_api_credentials("api.txt")
			client = OpenAI(api_key=api_key, base_url=base_url)
			
			# Build index for the new file
			print(f"Processing new source file: {source_file}")
			embeddings, chunks, index_id = build_or_load_index(
				corpus_path=source_file,
				embedding_model="text-embedding-3-large",
				client=client,
				chunk_size_chars=8000,
				overlap_chars=1000,
			)
			
			# Refresh the cache list
			nonlocal caches
			caches = list_available_caches()
			
			# Clear and repopulate treeview
			for item in tree.get_children():
				tree.delete(item)
			
			for cache in caches:
				tree.insert("", tk.END, values=(
					cache['original_name'],
					f"{cache['total_size_mb']:.1f}",
					cache['cache_id']
				))
			
			# Select all by default
			for item in tree.get_children():
				tree.selection_add(item)
			
			# Update total size
			total_size = sum(cache['total_size_mb'] for cache in caches)
			size_label.config(text=f"Total cache size: {total_size:.1f} MB")
			
			messagebox.showinfo("Success", f"Successfully added source file: {os.path.basename(source_file)}")
			
		except Exception as e:
			messagebox.showerror("Error", f"Failed to add source file: {str(e)}")
	
	def on_add_multiple_source_files():
		"""Add multiple source files and refresh the cache list."""
		# Select multiple source files
		source_files = filedialog.askopenfilenames(
			title="Select multiple source files (TXT, PDF, Word, Excel)",
			filetypes=[
				("All supported files", "*.txt *.pdf *.docx *.doc *.xlsx *.xls"),
				("Text files", "*.txt"),
				("PDF files", "*.pdf"),
				("Word documents", "*.docx *.doc"),
				("Excel files", "*.xlsx *.xls"),
				("All files", "*.*")
			]
		)
		
		if not source_files:
			return
		
		try:
			# Load API credentials
			api_key, base_url = load_api_credentials("api.txt")
			client = OpenAI(api_key=api_key, base_url=base_url)
			
			# Process each file
			success_count = 0
			total_files = len(source_files)
			
			for i, source_file in enumerate(source_files):
				try:
					print(f"Processing file {i+1}/{total_files}: {source_file}")
					embeddings, chunks, index_id = build_or_load_index(
						corpus_path=source_file,
						embedding_model="text-embedding-3-large",
						client=client,
						chunk_size_chars=8000,
						overlap_chars=1000,
					)
					success_count += 1
				except Exception as e:
					print(f"Failed to process {source_file}: {str(e)}")
					messagebox.showerror("Error", f"Failed to process {os.path.basename(source_file)}: {str(e)}")
			
			# Refresh the cache list
			nonlocal caches
			caches = list_available_caches()
			
			# Clear and repopulate treeview
			for item in tree.get_children():
				tree.delete(item)
			
			for cache in caches:
				tree.insert("", tk.END, values=(
					cache['original_name'],
					f"{cache['total_size_mb']:.1f}",
					cache['cache_id']
				))
			
			# Select all by default
			for item in tree.get_children():
				tree.selection_add(item)
			
			# Update total size
			total_size = sum(cache['total_size_mb'] for cache in caches)
			size_label.config(text=f"Total cache size: {total_size:.1f} MB")
			
			messagebox.showinfo("Success", f"Successfully processed {success_count}/{total_files} source files")
			
		except Exception as e:
			messagebox.showerror("Error", f"Failed to process source files: {str(e)}")
	
	def on_select_excel():
		"""Select Excel file with paragraphs for verification."""
		nonlocal selected_statements_file
		statements_file = filedialog.askopenfilename(
			title="Select Excel file with paragraphs to verify",
			filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
		)
		
		if statements_file:
			selected_statements_file = statements_file
			messagebox.showinfo("Success", f"Selected statements file: {os.path.basename(statements_file)}")
	
	# Bind selection change event
	tree.bind('<<TreeviewSelect>>', lambda e: update_total_size())
	
	ttk.Button(button_frame, text="Statements to check", command=on_select_excel).grid(row=0, column=0, padx=(0, 5))
	ttk.Button(button_frame, text="Add Source files", command=on_add_multiple_source_files).grid(row=0, column=1, padx=(0, 5))
	ttk.Button(button_frame, text="Select All", command=on_select_all).grid(row=0, column=2, padx=(0, 5))
	ttk.Button(button_frame, text="Deselect All", command=on_deselect_all).grid(row=0, column=3, padx=(0, 10))
	ttk.Button(button_frame, text="Delete Selected", command=on_delete_selected_caches).grid(row=0, column=4, padx=(0, 10))
	ttk.Button(button_frame, text="Let's check!", command=on_ok).grid(row=0, column=5)
	ttk.Button(button_frame, text="Exit", command=on_cancel).grid(row=0, column=6)
	
	# Configure grid weights
	root.columnconfigure(0, weight=1)
	root.rowconfigure(0, weight=1)
	main_frame.columnconfigure(0, weight=1)
	main_frame.rowconfigure(1, weight=1)
	tree_frame.columnconfigure(0, weight=1)
	tree_frame.rowconfigure(0, weight=1)
	
	# Initialize total size display
	update_total_size()
	
	root.mainloop()
	return selected_caches, selected_statements_file

def select_caches_cli() -> List[str]:
	"""Command line interface to select caches."""
	caches = list_available_caches()
	if not caches:
		print("No cached files found. Please run the script with a source file first to create caches.")
		return []
	
	print("\nAvailable cached source files:")
	print(f"{'#':<3} {'Source File':<40} {'Size (MB)':<10} {'Cache ID':<20}")
	print("-" * 75)
	
	total_size = 0
	for i, cache in enumerate(caches):
		size_str = f"{cache['total_size_mb']:.1f}"
		print(f"{i+1:<3} {cache['original_name']:<40} {size_str:<10} {cache['cache_id']:<20}")
		total_size += cache['total_size_mb']
	
	print("-" * 75)
	print(f"Total cache size: {total_size:.1f} MB")
	
	while True:
		try:
			selection = input("\nEnter cache numbers to use (comma-separated, e.g., 1,2,3) or 'all': ").strip()
			if selection.lower() == 'all':
				selected_caches = [cache["cache_id"] for cache in caches]
				selected_size = sum(cache['total_size_mb'] for cache in caches)
				print(f"Selected {len(selected_caches)} caches, total size: {selected_size:.1f} MB")
				return selected_caches
			
			indices = [int(x.strip()) - 1 for x in selection.split(",")]
			if all(0 <= i < len(caches) for i in indices):
				selected_caches = [caches[i]["cache_id"] for i in indices]
				selected_size = sum(caches[i]['total_size_mb'] for i in indices)
				print(f"Selected {len(selected_caches)} caches, total size: {selected_size:.1f} MB")
				return selected_caches
			else:
				print("Invalid selection. Please try again.")
		except (ValueError, IndexError):
			print("Invalid input. Please enter numbers separated by commas.")


def delete_cache(cache_id: str, cache_dir: str = ".index_cache") -> bool:
	"""Delete a specific cache and return success status."""
	npz_path = os.path.join(cache_dir, f"{cache_id}.npz")
	chunks_path = os.path.join(cache_dir, f"{cache_id}.chunks.jsonl")
	
	deleted_files = 0
	try:
		if os.path.exists(npz_path):
			os.remove(npz_path)
			deleted_files += 1
		if os.path.exists(chunks_path):
			os.remove(chunks_path)
			deleted_files += 1
		return deleted_files == 2  # Both files should be deleted
	except Exception as e:
		print(f"Error deleting cache {cache_id}: {str(e)}")
		return False

def delete_caches_gui() -> None:
	"""GUI to delete selected caches."""
	try:
		import tkinter as tk
		from tkinter import ttk, messagebox
	except ImportError:
		print("tkinter not available, using command line deletion")
		delete_caches_cli()
		return
	
	caches = list_available_caches()
	if not caches:
		messagebox.showwarning("No Caches", "No cached files found to delete.")
		return
	
	root = tk.Tk()
	root.title("Delete Source Caches")
	root.geometry("800x500")
	
	# Create main frame
	main_frame = ttk.Frame(root, padding="10")
	main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
	
	# Instructions
	ttk.Label(main_frame, text="Select caches to delete (this action cannot be undone):", font=("Arial", 10, "bold")).grid(row=0, column=0, columnspan=3, pady=(0, 10), sticky=tk.W)
	
	# Create treeview for better display
	tree_frame = ttk.Frame(main_frame)
	tree_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
	
	# Create treeview with columns
	columns = ("filename", "size", "id")
	tree = ttk.Treeview(tree_frame, columns=columns, show="tree headings", height=15)
	
	# Configure columns
	tree.heading("filename", text="Source File")
	tree.heading("size", text="Cache Size (MB)")
	tree.heading("id", text="Cache ID")
	
	tree.column("filename", width=300, anchor=tk.W)
	tree.column("size", width=100, anchor=tk.E)
	tree.column("id", width=150, anchor=tk.W)
	
	# Add scrollbar
	scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
	tree.configure(yscrollcommand=scrollbar.set)
	
	tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
	scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
	
	# Populate treeview
	for cache in caches:
		tree.insert("", tk.END, values=(
			cache['original_name'],
			f"{cache['total_size_mb']:.1f}",
			cache['cache_id']
		))
	
	# Total size display
	total_size = sum(cache['total_size_mb'] for cache in caches)
	size_label = ttk.Label(main_frame, text=f"Total cache size: {total_size:.1f} MB", font=("Arial", 9))
	size_label.grid(row=2, column=0, columnspan=2, pady=(5, 0), sticky=tk.W)
	
	# Buttons
	button_frame = ttk.Frame(main_frame)
	button_frame.grid(row=3, column=0, columnspan=3, pady=(10, 0))
	
	def update_total_size():
		"""Update total size display based on selection."""
		selection = tree.selection()
		selected_size = sum(
			cache['total_size_mb'] 
			for cache in caches 
			if cache['cache_id'] in [tree.item(item)['values'][2] for item in selection]
		)
		size_label.config(text=f"Selected for deletion: {selected_size:.1f} MB")
	
	def on_delete():
		selection = tree.selection()
		if not selection:
			messagebox.showwarning("No Selection", "Please select at least one cache to delete.")
			return
		
		selected_caches = [tree.item(item)['values'][2] for item in selection]
		selected_names = [tree.item(item)['values'][0] for item in selection] # Changed to original_name
		selected_size = sum(
			cache['total_size_mb'] 
			for cache in caches 
			if cache['cache_id'] in selected_caches
		)
		
		# Confirmation dialog
		confirm_text = f"Are you sure you want to delete {len(selected_caches)} cache(s)?\n\n"
		confirm_text += f"Files to delete:\n"
		for name in selected_names:
			confirm_text += f"• {name}\n"
		confirm_text += f"\nTotal size to free: {selected_size:.1f} MB\n\nThis action cannot be undone!"
		
		if not messagebox.askyesno("Confirm Deletion", confirm_text):
			return
		
		# Delete selected caches
		success_count = 0
		for cache_id in selected_caches:
			if delete_cache(cache_id):
				success_count += 1
		
		# Show results
		if success_count == len(selected_caches):
			messagebox.showinfo("Success", f"Successfully deleted {success_count} cache(s).")
		else:
			messagebox.showwarning("Partial Success", f"Deleted {success_count}/{len(selected_caches)} cache(s). Some files may have been in use.")
		
		root.destroy()
	
	def on_cancel():
		root.destroy()
	
	def on_select_all():
		for item in tree.get_children():
			tree.selection_add(item)
		update_total_size()
	
	def on_deselect_all():
		tree.selection_remove(tree.selection())
		update_total_size()
	
	# Bind selection change event
	tree.bind('<<TreeviewSelect>>', lambda e: update_total_size())
	
	ttk.Button(button_frame, text="Select All", command=on_select_all).grid(row=0, column=0, padx=(0, 5))
	ttk.Button(button_frame, text="Deselect All", command=on_deselect_all).grid(row=0, column=1, padx=(0, 10))
	ttk.Button(button_frame, text="Delete Selected", command=on_delete).grid(row=0, column=2, padx=(0, 10))
	ttk.Button(button_frame, text="Cancel", command=on_cancel).grid(row=0, column=3)
	
	# Configure grid weights
	root.columnconfigure(0, weight=1)
	root.rowconfigure(0, weight=1)
	main_frame.columnconfigure(0, weight=1)
	main_frame.rowconfigure(1, weight=1)
	tree_frame.columnconfigure(0, weight=1)
	tree_frame.rowconfigure(0, weight=1)
	
	# Initialize total size display
	update_total_size()
	
	root.mainloop()

def delete_caches_cli() -> None:
	"""Command line interface to delete caches."""
	caches = list_available_caches()
	if not caches:
		print("No cached files found to delete.")
		return
	
	print("\nAvailable cached source files:")
	print(f"{'#':<3} {'Source File':<40} {'Size (MB)':<10} {'Cache ID':<20}")
	print("-" * 75)
	
	total_size = 0
	for i, cache in enumerate(caches):
		size_str = f"{cache['total_size_mb']:.1f}"
		print(f"{i+1:<3} {cache['original_name']:<40} {size_str:<10} {cache['cache_id']:<20}")
		total_size += cache['total_size_mb']
	
	print("-" * 75)
	print(f"Total cache size: {total_size:.1f} MB")
	
	while True:
		try:
			selection = input("\nEnter cache numbers to delete (comma-separated, e.g., 1,2,3) or 'all': ").strip()
			if selection.lower() == 'all':
				selected_caches = [cache["cache_id"] for cache in caches]
				selected_names = [cache["original_name"] for cache in caches]
				selected_size = sum(cache['total_size_mb'] for cache in caches)
			else:
				indices = [int(x.strip()) - 1 for x in selection.split(",")]
				if not all(0 <= i < len(caches) for i in indices):
					print("Invalid selection. Please try again.")
					continue
				
				selected_caches = [caches[i]["cache_id"] for i in indices]
				selected_names = [caches[i]["original_name"] for i in indices]
				selected_size = sum(caches[i]['total_size_mb'] for i in indices)
			
			# Confirmation
			print(f"\nYou are about to delete {len(selected_caches)} cache(s):")
			for name in selected_names:
				print(f"• {name}")
			print(f"Total size to free: {selected_size:.1f} MB")
			
			confirm = input("\nAre you sure? This action cannot be undone! (yes/no): ").strip().lower()
			if confirm not in ['yes', 'y']:
				print("Deletion cancelled.")
				return
			
			# Delete caches
			success_count = 0
			for cache_id in selected_caches:
				if delete_cache(cache_id):
					success_count += 1
					print(f"✓ Deleted cache: {cache_id}")
				else:
					print(f"✗ Failed to delete cache: {cache_id}")
			
			print(f"\nDeletion complete: {success_count}/{len(selected_caches)} cache(s) deleted successfully.")
			return
			
		except (ValueError, IndexError):
			print("Invalid input. Please enter numbers separated by commas.")


def process_statement_batch(client: OpenAI, judge_model: str, embedding_model: str, 
						   statements_batch: List, doc_matrix_norm: np.ndarray, 
						   chunks: List[Tuple[int, str]], top_k: int, max_snippet_chars: int,
						   is_court_format: bool) -> List[Dict]:
	"""Process a batch of statements in parallel."""
	results = []
	
	# Batch embed all statements in the batch
	statements_texts = []
	for stmt in statements_batch:
		if is_court_format:
			# Get three-line context for court format
			idx = stmt['index']
			three_lines = []
			if idx > 0:
				three_lines.append(statements_batch[idx-1]['content'] if idx-1 < len(statements_batch) else "")
			three_lines.append(stmt['content'])
			if idx < len(statements_batch) - 1:
				three_lines.append(statements_batch[idx+1]['content'] if idx+1 < len(statements_batch) else "")
			statements_texts.append("\n".join(three_lines))
		else:
			statements_texts.append(stmt)
	
	# Batch embed
	batch_embeddings = embed_texts(client, embedding_model, statements_texts)
	
	# Process each statement in the batch with individual progress
	for i, stmt in enumerate(statements_batch):
		# Retrieve evidence
		q_vec = batch_embeddings[i]
		indices = top_k_indices_cosine(doc_matrix_norm, q_vec, top_k)
		evidence = []
		for idx in indices:
			text = chunks[idx][1]
			if len(text) > max_snippet_chars:
				text = text[:max_snippet_chars] + "..."
			evidence.append(text)
		
		# Judge with model
		if is_court_format:
			result = judge_court_statement(client, judge_model, stmt['par'], statements_texts[i], evidence)
		else:
			result = judge_statement(client, judge_model, stmt, evidence)
		
		results.append({
			'statement': stmt,
			'result': result,
			'evidence': evidence
		})
	
	return results


def process_statements_parallel(client: OpenAI, judge_model: str, embedding_model: str,
							   statements: List, doc_matrix_norm: np.ndarray,
							   chunks: List[Tuple[int, str]], top_k: int, max_snippet_chars: int,
							   is_court_format: bool, max_workers: int = 8) -> List[Dict]:  # Changed to 8
	"""Process statements in parallel using ThreadPoolExecutor."""
	all_results = []
	batch_size = 10  # Reduced from 25 to 10
	
	# Split statements into batches
	batches = []
	for i in range(0, len(statements), batch_size):
		batch = statements[i:i + batch_size]
		batches.append(batch)
	
	print(f"Processing {len(statements)} statements in {len(batches)} batches of {batch_size}...")
	
	# Process batches in parallel with detailed progress
	with ThreadPoolExecutor(max_workers=max_workers) as executor:
		futures = []
		for batch in batches:
			future = executor.submit(
				process_statement_batch,
				client, judge_model, embedding_model, batch,
				doc_matrix_norm, chunks, top_k, max_snippet_chars, is_court_format
			)
			futures.append(future)
		
		# Collect results with detailed progress bar
		completed_batches = 0
		for future in tqdm(as_completed(futures), total=len(futures), 
						  desc="Processing statements", unit="batch"):
			batch_results = future.result()
			all_results.extend(batch_results)
			completed_batches += 1
			
			# Show intermediate progress
			completed_statements = len(all_results)
			if completed_batches % 2 == 0 or completed_batches == len(batches):  # Update every 2 batches
				print(f"  Completed {completed_statements}/{len(statements)} statements ({completed_statements/len(statements)*100:.1f}%)")
	
	return all_results


def run(
	api_file: str,
	corpus_path: str,
	statements_path: str,
	output_excel: str,
	checkpoint_path: str,
	judge_model: str,
	embedding_model: str,
	chunk_size_chars: int,
	overlap_chars: int,
	top_k: int,
	max_snippet_chars: int,
	use_caches: bool = False,
) -> None:
	# Global variable to track current Excel path
	global current_excel_path
	current_excel_path = output_excel
	api_key, base_url = load_api_credentials(api_file)
	client = OpenAI(api_key=api_key, base_url=base_url)

	# Load embeddings and chunks
	if use_caches:
		# Use multiple caches
		selected_cache_ids, gui_statements_path = select_caches_gui()
		if not selected_cache_ids:
			print("No caches selected. Exiting.")
			return
		
		# Use statements file from GUI if provided
		if gui_statements_path:
			statements_path = gui_statements_path
		else:
			print("No statements file selected. Exiting.")
			return
		
		print(f"Loading {len(selected_cache_ids)} caches...")
		embeddings, chunks, cache_names = load_multiple_caches(selected_cache_ids)
		print(f"Loaded caches: {', '.join(cache_names)}")
		print(f"Total chunks: {len(chunks)}")
	else:
		# Build or load single index
		embeddings, chunks, index_id = build_or_load_index(
			corpus_path=corpus_path,
			embedding_model=embedding_model,
			client=client,
			chunk_size_chars=chunk_size_chars,
			overlap_chars=overlap_chars,
		)
	
	# Check if statements file exists
	if not os.path.exists(statements_path):
		raise RuntimeError(f"Statements file not found: {statements_path}")
	
	statements = read_statements(statements_path)
	if not statements:
		raise RuntimeError("No statements to process")

	# Check if this is court judgment format (list of dicts)
	is_court_format = isinstance(statements[0], dict) if statements else False
	
	doc_matrix_norm = normalize_matrix(embeddings)

	if is_court_format:
		# Court judgment analysis with parallel processing
		output_path, _ = excel_prepare_court_writer(output_excel)
		current_excel_path = output_path  # Update global path
		checkpoint = load_court_checkpoint(checkpoint_path)
		processed = set(checkpoint.keys())

		# Filter out already processed statements
		unprocessed_statements = [stmt for stmt in statements if stmt['par'] not in processed]
		
		if not unprocessed_statements:
			print("All statements already processed.")
			return
		
		print(f"Processing {len(unprocessed_statements)} statements in parallel...")
		
		# Process statements in parallel
		results = process_statements_parallel(
			client, judge_model, embedding_model, unprocessed_statements,
			doc_matrix_norm, chunks, top_k, max_snippet_chars, is_court_format,
			max_workers=8  # or 12 depending on your system
		)
		
		# Write results to Excel and checkpoint
		print(f"Writing {len(results)} results to Excel and checkpoint...")
		for i, result in enumerate(tqdm(results, desc="Writing results", unit="result")):
			stmt = result['statement']
			verdict = result['result'].get("verdict", "not accurate")
			degree_of_accuracy = result['result'].get("degree_of_accuracy", 5)
			inaccuracy_type = result['result'].get("inaccuracy_type", "none")
			description = result['result'].get("description", "")
			
			# Include the original paragraph content
			par_context = stmt['content']
			
			excel_append_row(current_excel_path, [stmt['par'], par_context, verdict, degree_of_accuracy, inaccuracy_type, description])
			append_court_checkpoint(checkpoint_path, stmt['par'], par_context, verdict, degree_of_accuracy, inaccuracy_type, description)
		
		print(f"Processed {len(results)} statements in parallel.")
		
		# Sort the Excel file by paragraph number
		sort_excel_by_paragraph_number(current_excel_path)
		
	else:
		# Standard analysis with parallel processing
		headers = ["index", "statement", "verdict", "explanation"]
		_, _ = excel_prepare_writer(output_excel, headers)
		checkpoint = load_checkpoint(checkpoint_path)

		processed = set(checkpoint.keys())

		# Filter out already processed statements
		unprocessed_statements = [stmt for i, stmt in enumerate(statements) if i not in processed]
		
		if not unprocessed_statements:
			print("All statements already processed.")
			return
		
		print(f"Processing {len(unprocessed_statements)} statements in parallel...")
		
		# Process statements in parallel
		results = process_statements_parallel(
			client, judge_model, embedding_model, unprocessed_statements,
			doc_matrix_norm, chunks, top_k, max_snippet_chars, is_court_format,
			max_workers=8  # or 12 depending on your system
		)
		
		# Write results to Excel and checkpoint
		print(f"Writing {len(results)} results to Excel and checkpoint...")
		for i, result in enumerate(tqdm(results, desc="Writing results", unit="result")):
			stmt = result['statement']
			verdict = result['result'].get("verdict", "not accurate")
			explanation = result['result'].get("explanation", "")
			
			# Find original index
			original_index = statements.index(stmt)
			
			excel_append_row(output_excel, [original_index, stmt, verdict, explanation])
			append_checkpoint(checkpoint_path, original_index, stmt, verdict, explanation)
		
		print(f"Processed {len(results)} statements in parallel.")

	print(f"Done. Results written to: {current_excel_path}\nCheckpoint: {checkpoint_path}")


def default_paths_in_workspace() -> Tuple[str, str, str]:
	# Heuristics for defaults in current workspace
	corpus = None
	statements = None
	for name in os.listdir("."):
		low = name.lower()
		if low.endswith(".txt") and "hearing" in low:
			corpus = name
		if (low.endswith(".txt") and ("ws" in low or "statements" in low)) or low.endswith(".xlsx"):
			if statements is None:
				statements = name
	if corpus is None:
		for name in os.listdir("."):
			if name.lower().endswith(".txt") and name != "api.txt":
				corpus = name
				break
	if statements is None:
		# Prefer Excel if present
		for name in os.listdir("."):
			if name.lower().endswith(".xlsx"):
				statements = name
				break
	return corpus or "Hearing Transcripts ALL.txt", statements or "RozhkovaWS.txt", "verification_results.xlsx"


def maybe_pick_files_gui(args) -> None:
	try:
		import tkinter as tk
		from tkinter import filedialog, messagebox
	except Exception:
		return
	root = tk.Tk()
	root.withdraw()
	try:
		# Always use caches and go to cache selection GUI
		args.use_caches = True
		
		# Don't ask for statements file - it will be handled in cache selection GUI
		# Don't ask for output file - use default verification_results.xlsx
		args.output_excel = "verification_results.xlsx"

		# Don't ask for checkpoint file - use default
		args.checkpoint_path = "verification_checkpoint.jsonl"
	finally:
		try:
			root.destroy()
		except Exception:
			pass


if __name__ == "__main__":
	parser = argparse.ArgumentParser(description="Verify statements against a large corpus with retrieval + LLM")
	parser.add_argument("--api", dest="api_file", default="api.txt", help="Path to API credentials file")
	# Inputs
	default_corpus, default_statements, default_output = default_paths_in_workspace()
	parser.add_argument("--corpus", default=default_corpus, help="Path to source file (TXT, PDF, Word, Excel)")
	parser.add_argument("--statements", default=default_statements, help="Path to statements (.txt lines or .xlsx column)")
	# Outputs
	parser.add_argument("--out", dest="output_excel", default=default_output, help="Output Excel file path")
	parser.add_argument("--ckpt", dest="checkpoint_path", default="verification_checkpoint.jsonl", help="Checkpoint file path")
	# Models
	parser.add_argument("--judge-model", default="gpt-5-mini", help="Model for verification")
	parser.add_argument("--embed-model", default="text-embedding-3-large", help="Model for embeddings")
	# Retrieval
	parser.add_argument("--chunk-chars", type=int, default=4000, help="Chunk size in characters for corpus")
	parser.add_argument("--overlap-chars", type=int, default=500, help="Overlap in characters between chunks")
	parser.add_argument("--top-k", type=int, default=10, help="Top-k evidence chunks to retrieve")
	parser.add_argument("--max-snippet-chars", type=int, default=800, help="Max characters per evidence snippet")
	# GUI picker
	parser.add_argument("--gui", action="store_true", help="Open Windows dialogs to pick files (API, corpus, statements, output, checkpoint)")
	# Cache selection
	parser.add_argument("--use-caches", action="store_true", help="Use existing cached source files instead of a new source file")
	# Cache management
	parser.add_argument("--delete-caches", action="store_true", help="Delete selected cached source files")

	args = parser.parse_args()

	# Handle cache deletion
	if args.delete_caches:
		delete_caches_gui()
	else:
		if args.gui:
			maybe_pick_files_gui(args)

		run(
			api_file=args.api_file,
			corpus_path=args.corpus,
			statements_path=args.statements,
			output_excel=args.output_excel,
			checkpoint_path=args.checkpoint_path,
			judge_model=args.judge_model,
			embedding_model=args.embed_model,
			chunk_size_chars=args.chunk_chars,
			overlap_chars=args.overlap_chars,
			top_k=args.top_k,
			max_snippet_chars=args.max_snippet_chars,
			use_caches=args.use_caches,
		)
